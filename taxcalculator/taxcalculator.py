import pdfplumber
from openpyxl import load_workbook
import fitz
import os
from PIL import Image, ImageDraw, ImageFont
"""
cells for input 

In XLXS file:
1a == D5
1b == C6
2a == D7
2b == D8
2c == D9
2d == D10
2e == NOT INCLUDED
2f == NOT INCLUDED
3-15 == D11-D23

"""
cellpostions = ["D5", "C6" , "D7", "D8", "D9", "D10", "NIL", "NIL", "D11", "D12", "D13", "D14", "D15", "D16", "D17", "D18", "D19", "D20", "D21", "D22", "D23" ]


#THIS FUNCTION IS FOR TURNING NUMBERS INTO READABLE AMOUNTS e.g. (11111 > 11,111)
def readable(num):
    num = list(str(num))
    if len(num) >= 4:
        num.insert(len(num) % 3, ",")

    return "".join(num)


dic = [                           #keys are unimportant, the values are the 'bounding boxes' for extracting the text
    ['1', [646, 293, 783, 302]], 
    ['1', [556, 313, 693, 322]], 
    ['0.00', [666, 322, 782, 331]], 
    ['0.00', [576, 342, 692, 351]], 
    ['0.00', [576, 352, 692, 361]], 
    ['0.00', [576, 362, 692, 371]], 
    ['0.00', [576, 372, 692, 381]], 
    ['0.00', [576, 381, 692, 390]], 
    ['0.00', [666, 391, 782, 400]], 
    ['0.00', [666, 401, 782, 410]], 
    ['1', [567, 411, 692, 420]], 
    ['0.00', [666, 421, 782, 430]],
    ['0.00', [666, 431, 782, 440]], 
    ['0.00', [666, 450, 782, 459]], 
    ['0.00', [666, 460, 782, 469]], 
    ['0.00', [666, 470, 782, 479]], 
    ['0.00', [576, 490, 692, 499]], 
    ['0.00', [666, 519, 782, 528]]
]
extractednums = []
textextractednums = []

#<----------------------------- GET FILEPATHS ----------------------------->
for filename in os.listdir("DROPHERE/"):
    if "1099" in filename and ".pdf" in filename:
        pdffilepath = "DROPHERE/" + filename
    elif ".xlsx" in filename and "1099neeewesr" in filename:    
        worksheetfilepath = "DROPHERE/" + filename
    elif "1041" in filename and ".pdf" in filename:
        pdftowritefilepath = "DROPHERE/1041.pdf"    

##<----------------------------- GET NUMBERS FROM PDF ----------------------------->
#here i extract the numbers with bounding boxes and remove '$' and ',' so i can turn them to ints
with pdfplumber.open(pdffilepath) as pdf:
    page = pdf.pages[2] 
    for i in range(len(dic)):
        x = page.within_bbox((dic[i][1][0], dic[i][1][1], dic[i][1][2], dic[i][1][3])).extract_text()
        x = x.replace("$", "")
        x = x.replace(",", "")
        x = x.strip()
        print(x)
        extractednums.append(x)


# compare boundingbox extract with text extract (i only found out there was a simpler way to extract text without the bounding boxes after i created the bounding boxes )
# i didnt want to waste my work, so i just compared with this, where i used the .extract_text() function and used the for loop to get the numbers from the text
lst = list(page.extract_text())
adding = True
x = ""
for i in range(len(lst)):
    if lst[i] == "$":
        adding = True
    elif lst[i].isalpha() or lst[i] == '\n':
        adding = False
        if x != '':
            textextractednums.append(x)
        x = ''    
    if adding and (lst[i].isdigit() or lst[i] == '.'):
        x += lst[i]
if extractednums != textextractednums:
    print("ERROR IN EXTRACTING TEXT FROM PDF")

##<----------------------------- ADD TO CALCULATER SHEET ----------------------------->

#add to calculator sheet
xlsheet = load_workbook(worksheetfilepath)
print(xlsheet)
for i in range(len(dic)):
    if cellpostions[i] == "NIL":
        continue
    xlsheet["XXXX-3380.csv"][cellpostions[i]] = float(extractednums[i])
    
xlsheet.save(worksheetfilepath)

print(textextractednums)
print("OPEN THE CALCULATOR WORKSHEET AND SAVE.")
input("PRESS ENTER WHEN DONE ")
# ^^ i had to do this otherwise the cells wouldnt return the newly calculated values


#extract CALCULATED values from sheet
"""
#DIMENSONS OF PDF IS 6800 x 8800 pixels at 800DPU
#FONT COLOR IS 0 0 127
#distance between each is ~130 boxes starting from 3:2820

PAGE 1 COORDINATES OF BOXES TO WRITE:
    1:2430
   2a:2560
    3:2820
    9:3630
   21:5500
   22:5630
   23:5760
   24:5900
   26:6160
   27:6290
   28:6430

PAGE 2 COORDINATES OF BOXES TO WRITE (only on G):
    1a:5000, 5500
    1d:5500
    9:6960

PAGE 3 COORDINATES OF BOXES TO WRITE:
    '10': 830
    '19': 2160
"""
dic0 = {
    'Name': ['(taken out)', (1530, 960)],
    'Name2': ['(taken out)', (1530, 1230)],
    'Street': ['(taken out)', (1530, 1640)],
    'City': ['(taken out)', (1530, 1900)],
    'ID': ['(taken out)', (5401, 960)],
    'Date': ['(taken out)', (5400, 1230)]
}

dic1 = {
    '1': [(6120, 2430)],
    '2a': [(6120, 2560)],
    '2b': [(6120, 2690)],
    '3': [(6120, 2820)],
    '9': [(6120, 3630)],
    '21': [(6120, 5500)],
    '22': [(6120, 5630)],
    '23': [(6120, 5760)],
    '24': [(6120, 5900)],
    '26': [(6120, 6160)],
    '27': [(6120, 6290)],
    '28': [(6120, 6430)]
}
dic2 = {
    '1a': [(5080, 4700)],
    '1d': [(6120, 5100)],
    '9': [(6120, 6960)],
}

dic3 = {
    '10': [(6120, 830)],
    '12': [(6120, 1090)],
    '19': [(6120, 2160)], 
    'box1': [(6120, 2420)], 
    'box2': [(6120, 2810)], 
    'box3': [(6120, 3070)], 
    'box4': [(6120, 3610)], 
    'box5': [(6120, 3870)], 
    'box6': [(6120, 4550)], 
    'box7': [(6120, 4810)], 
    'box8': [(6120, 4960)], 
    'box9': [(6120, 5350)], 
    'box10': [(6120, 5740)]
}

calculatednums = []
xlsheetcalculated = load_workbook(worksheetfilepath, data_only=True)
##<----------------------------- EXTRACT FROM CALCULATER SHEET ----------------------------->

#page1
dic1['1'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=15, column=37).value)#1
dic1['2a'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=16, column=37).value)#2a
dic1['2b'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=17, column=37).value)#2b
dic1['3'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=18, column=37).value)#3
dic1['9'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=24, column=37).value)#9
dic1['21'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=38, column=37).value)#21
dic1['22'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=39, column=37).value)#22
dic1['23'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=41, column=37).value)#23
dic1['24'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=42, column=37).value)#24
dic1['26'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=44, column=37).value)#26
dic1['27'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=45, column=37).value)#27
dic1['28'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=46, column=37).value)#28
#page2
dic2['1a'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=80, column=37).value)#1a
dic2['1d'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=84, column=37).value)#1d
dic2['9'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=98, column=37).value)#9
#page3
dic3['10'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=104, column=37).value)#10
dic3['12'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=106, column=37).value)#12
dic3['19'].insert(0, xlsheetcalculated["XXXX-3380.csv"].cell(row=115, column=37).value)#19

print(len(calculatednums), calculatednums)
print(dic1)
print(dic2)
print(dic3)



#<----------------------------- WRITE ONTO PNG  ----------------------------->

doc = fitz.open(pdftowritefilepath)           
page1 = doc[0] 
page2 = doc[1]
page3 = doc[2]

pixelmap1 = page1.get_pixmap(dpi=800)
pixelmap2 = page2.get_pixmap(dpi=800)
pixelmap3 = page3.get_pixmap(dpi=800)
#^^ here i use fitz to turn each page of the 1041 pdf to pngs

img1 = Image.frombytes("RGB", [pixelmap1.width, pixelmap1.height], pixelmap1.samples)
img2 = Image.frombytes("RGB", [pixelmap2.width, pixelmap2.height], pixelmap2.samples)
img3 = Image.frombytes("RGB", [pixelmap3.width, pixelmap3.height], pixelmap3.samples)

draw1 = ImageDraw.Draw(img1)
draw2 = ImageDraw.Draw(img2)
draw3 = ImageDraw.Draw(img3)

font = ImageFont.truetype("arial", 90)

for i in dic0.values():
    draw1.text((i[1][0], i[1][1]), readable(i[0]), font=font, fill=(0, 0, 127))


for i in dic1.values():
    if i[0] == None:
        continue
    print(readable(i[0]))
    draw1.text((i[1][0]+((5 - len(str(i[0])))*50), i[1][1]), readable(i[0]), font=font, fill=(0, 0, 127))

for i in dic2.values():
    if i[0] == None:
        continue
    print(readable(i[0]))
    draw2.text((i[1][0]+((5 - len(str(i[0])))*50), i[1][1]), readable(i[0]), font=font, fill=(0, 0, 127))

for i in dic3.values():
    if i[0] == None:
        continue
    if len(i) < 2:
        continue 
    print(readable(i[0]), )
    if i[0] == 'X':
        draw3.text((i[1][0]+135, i[1][1]), readable(i[0]), font=font, fill=(0, 0, 127))
    else:
        draw3.text((i[1][0]+((5 - len(str(i[0])))*50), i[1][1]), readable(i[0]), font=font, fill=(0, 0, 127))






img1.save("written_image1.png")
img2.save("written_image2.png")
img3.save("written_image3.png")
print("END")
