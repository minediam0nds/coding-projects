from openpyxl import load_workbook
from openpyxl import Workbook
nameoffile = input("Enter name of file(excluding ' .xlsx')")
filepath = nameoffile + ".xlsx"
wb = load_workbook(filepath)
ws = wb["Sheet1"]
secondwb = Workbook()
writews = secondwb["Sheet"]


def columntoletter(num):
    if not 1 <= num <= 26:
        raise ValueError("Column number must be between 1 and 26")
    return chr(64 + num)

      

cell_values = []
writerow = 1
writecolumn = 1
row = int(input("Enter line which first stock name appears: "))
lastrow = int(input("Enter line which stock details end: "))
count = 0


print("++++++++++++")
while True:
    if row >= lastrow:
        print("STOPPED AT LINE ", row, ": LAST ROW REACHED")
        break

    cell = ws.cell(row=row, column=1)

    if (cell.alignment.horizontal == "right" and cell.value == "...") or (cell.alignment.horizontal == None and "%" in cell.value) or (cell.font and cell.font.color and cell.font.color.rgb == "FFFFFF"):
          print(f"CELL {"A"+str(row)} marked as unneeded {cell.value}, {cell.alignment.horizontal},  {cell.font.color.rgb}")
          row += 1

          continue
    

    elif (cell.alignment.horizontal == None or cell.alignment.horizontal == "left") and "." in cell.value and len(cell.value) > 10 and cell.value.upper() == cell.value:
          count += 1

          writerow += 1
          writecolumn = 1

          splitname = cell.value.split(".")
          writews[columntoletter(writecolumn)+str(writerow)] = splitname[0]+"."+splitname[-1][0:2]

          writecolumn += 1
          writews[columntoletter(writecolumn)+str(writerow)]= splitname[-1][2:]
          writecolumn += 1
          row += 1


          print(f"FOUND NAME: ADDED {cell.value} TO {columntoletter(writecolumn-1)+str(writerow)} and {columntoletter(writecolumn)+str(writerow)}")
          continue

    elif (cell.alignment.horizontal == None or cell.alignment.horizontal == "left") and "Total" in cell.value:
         row += 1
         

         while True:
              cell = ws.cell(row=row, column=1)
              if (cell.alignment.horizontal == None or cell.alignment.horizontal == "left") and "." in cell.value and len(cell.value) > 10 and cell.value.upper() == cell.value:
                   break
              if row >= lastrow:
                   break
              print(f"SKIPPED ROW {"A"+str(row)}, VALUE '{cell.value}': PART OF 'Total' Block")
              row += 1
         continue

              


    writews[columntoletter(writecolumn)+str(writerow)] = cell.value
    print(f"ADDED CELL {"A"+str(row)} with value '{cell.value}' to {columntoletter(writecolumn)+str(writerow)}")
    writecolumn += 1
    row += 1
    if cell.value == None:
        print("STOPPED AT LINE ", row, ": LINE IS EMPTY")
        break
    

    
secondwb.save("output.xlsx")
print("COMPLETE: COUNT = ", count, " SAVED AS output.xlsx")